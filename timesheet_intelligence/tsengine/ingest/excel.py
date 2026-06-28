"""Excel extractor (openpyxl). Produces one RawTable per sheet plus a text dump.

It does **no** interpretation -- header detection, weekly-vs-daily, time-in/out
maths all happen in the normalizer. Here we just faithfully dump the grid with
cell-level source references so any value remains auditable.
"""
from __future__ import annotations

import datetime as dt
import re
import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from ..schema import (ExtractionQuality, FileKind, RawExtraction, RawTable,
                      SourceRef)

_SOFFICE = [
    "soffice", "libreoffice", "/opt/homebrew/bin/soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice", "/usr/bin/libreoffice",
]


def _find_soffice() -> str | None:
    for c in _SOFFICE:
        p = shutil.which(c) if "/" not in c else (c if Path(c).exists() else None)
        if p:
            return p
    return None


def _to_xlsx(p: Path) -> Path:
    """openpyxl can't read legacy .xls -- convert to .xlsx via LibreOffice."""
    soffice = _find_soffice()
    if not soffice:
        return p
    tmp = Path(tempfile.mkdtemp(prefix="ts_xls_"))
    try:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir",
             str(tmp), str(p)],
            capture_output=True, timeout=120,
            env={"HOME": str(tmp), "PATH": "/usr/bin:/bin:/opt/homebrew/bin"},
        )
    except Exception:
        return p
    out = tmp / (p.stem + ".xlsx")
    return out if out.exists() and out.stat().st_size > 0 else p


def _cell_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        # midnight datetimes are really dates in these sheets
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.date().isoformat()
        return v.isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, dt.time):
        return v.strftime("%H:%M:%S")
    return v


def _trim_row(row: list[Any]) -> list[Any]:
    last = -1
    for i, c in enumerate(row):
        if c not in (None, ""):
            last = i
    return row[: last + 1]


def extract(path: str | Path) -> RawExtraction:
    p = Path(path)
    rel = p.name
    raw = RawExtraction(file=rel, kind=FileKind.EXCEL, quality=ExtractionQuality.NATIVE)
    load_path = p
    if p.suffix.lower() == ".xls":           # legacy format -> convert first
        load_path = _to_xlsx(p)
    try:
        # NOTE: read_only=True makes max_row/max_column unreliable on sparse
        # sheets (e.g. templates padded to 1000 rows), so we load normally.
        wb = openpyxl.load_workbook(str(load_path), data_only=True)
    except Exception as exc:
        # last-ditch: convert via LibreOffice even if the ext wasn't .xls
        if load_path == p:
            alt = _to_xlsx(p)
            if alt != p:
                try:
                    wb = openpyxl.load_workbook(str(alt), data_only=True)
                except Exception as exc2:
                    raw.notes.append(f"openpyxl failed after convert: {exc2}")
                    return raw
            else:
                raw.notes.append(f"openpyxl failed: {exc}")
                return raw
        else:
            raw.notes.append(f"openpyxl failed: {exc}")
            return raw

    text_lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue
        # cap absurd sheets (templates pad to 1000+ rows / wide column ranges)
        row_cap = min(max_row, 400)
        max_col = min(max_col, 64)
        grid: list[list[Any]] = []
        nonempty_rows = 0
        for r in range(1, row_cap + 1):
            row_vals: list[Any] = []
            for c in range(1, max_col + 1):
                row_vals.append(_cell_value(ws.cell(r, c).value))
            trimmed = _trim_row(row_vals)
            if any(v not in (None, "") for v in trimmed):
                nonempty_rows += 1
                grid.append(row_vals)
                text_lines.append(
                    f"[{sheet_name}!r{r}] " +
                    " | ".join("" if v is None else str(v) for v in trimmed)
                )
            else:
                grid.append(row_vals)  # keep blanks for positional alignment

        # strip fully-trailing empty rows
        while grid and all(v in (None, "") for v in grid[-1]):
            grid.pop()
        if not grid:
            continue

        table = RawTable(
            headers=[],  # header row is detected later; the grid is positional
            rows=grid,
            title=sheet_name,
            source=SourceRef(file=rel, sheet=sheet_name, extractor="excel"),
        )
        raw.tables.append(table)
        raw.sources.append(SourceRef(file=rel, sheet=sheet_name, extractor="excel"))

    raw.text = "\n".join(text_lines)
    raw.meta["sheet_names"] = wb.sheetnames
    raw.meta["name_hint"] = _name_hint(p.stem)
    if not raw.tables:
        raw.quality = ExtractionQuality.EMPTY
        raw.notes.append("no non-empty sheets")
    try:
        wb.close()
    except Exception:
        pass
    return raw


_MONTH_RE = (r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec|january|"
             r"february|march|april|june|july|august|september|october|november|"
             r"december)\b")

# Words that must never stand in as a person's name (months + generic timesheet
# vocabulary). Guards against filenames like 'May_timesheet_2026' -> "May".
_NAME_STOPWORDS = set(_MONTH_RE and [
    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "sept",
    "oct", "nov", "dec", "january", "february", "march", "april", "june",
    "july", "august", "september", "october", "november", "december",
    "timesheet", "time sheet", "timesheets", "monthly timesheet", "monthly",
    "time", "sheet", "ts", "month", "approved", "signed", "final", "copy",
])

# Employer / agency names that prefix some filenames (e.g. 'Ajace_Time Sheet_
# Justin_Thason') and must never be taken as the employee.
_COMPANY_STOPWORDS = {
    "ajace", "hcpss", "hexaware", "innososft", "innosoft", "npo", "hex",
}
# Leading tokens stripped so 'Company TimeSheet Name' yields the PERSON.
_LEAD_STRIP = ("time sheet ", "timesheet ", "timesheets ", "ts ")


def _name_hint(stem: str) -> str:
    """Best-effort employee name from a filename like 'Richard TS-April 2026'.

    Trims timesheet markers, month/year tails, and embedded date ranges so the
    registry can group files for the same person. The LLM/vision path refines
    the true name when available.
    """
    # underscores are filename separators -> normalize FIRST so month/marker
    # stripping sees real word boundaries (e.g. 'May_timesheet' not 'May_').
    s = stem.replace("_", " ")
    # Strip leading company names + timesheet markers so 'Company TimeSheet Name'
    # filenames yield the PERSON, not the company (e.g. 'Ajace Time Sheet Justin
    # Thason May-2026' -> 'Justin Thason', not 'Ajace').
    while True:
        sl = s.lstrip(" -·,")
        low2 = sl.lower()
        hit = next((p for p in _LEAD_STRIP if low2.startswith(p)), None) \
            or next((c + " " for c in _COMPANY_STOPWORDS if low2.startswith(c + " ")), None)
        if not hit:
            break
        s = sl[len(hit):]
    low = s.lower()
    cut = len(s)
    for mk in (" ts ", " ts-", "-ts ", " ts", "-ts",
               "timesheet", "time sheet", "timesheets"):
        i = low.find(mk)
        if i > 0:
            cut = min(cut, i)
    s = s[:cut]
    # strip embedded dates / numeric date ranges (e.g. 04-01-2026 to 04-30-2026)
    s = re.sub(r"\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}.*$", "", s)
    s = re.sub(r"\d{6,}.*$", "", s)              # e.g. 412026-4242026
    s = re.sub(_MONTH_RE + r".*$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\b20\d{2}\b", "", s)
    s = s.strip(" -·,")
    # never surface a bare month/company/generic word as a person's name; fall
    # back to the (unique) filename stem so records stay traceable and don't merge.
    if not s or s.lower() in _NAME_STOPWORDS or s.lower() in _COMPANY_STOPWORDS:
        return stem.replace("_", " ").strip(" -·,") or stem
    return s
